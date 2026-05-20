import argparse
import json
import re
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader


def normalize_header(value):
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())


def normalize_sku(value):
    return re.sub(r"\s+", "", str(value or "").strip().upper())


def clean_text(value):
    return " ".join(str(value or "").split()).strip()


def as_number(value):
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = str(value).strip().replace(",", "")
    if not cleaned:
        return 0.0
    try:
        return float(Decimal(cleaned))
    except (InvalidOperation, ValueError):
        return 0.0


def first_present(row, header_map, *names, default=""):
    for name in names:
        index = header_map.get(normalize_header(name))
        if index is None or index >= len(row):
            continue
        value = row[index]
        if value not in (None, ""):
            return clean_text(value)
    return default


def first_number(row, header_map, *names):
    for name in names:
        index = header_map.get(normalize_header(name))
        if index is None or index >= len(row):
            continue
        value = row[index]
        if value not in (None, ""):
            return as_number(value)
    return 0.0


def detect_header_row(rows, required_any, max_rows=12):
    for row_index, row in enumerate(rows[:max_rows], start=1):
        normalized = [normalize_header(cell) for cell in row]
        if any(key in normalized for key in required_any):
            return row_index, {key: idx for idx, key in enumerate(normalized) if key}
    return None, {}


def parse_item_master(file_path):
    suffix = Path(file_path).suffix.lower()
    if suffix == ".xls":
        raise ValueError("Legacy .xls item master files are not supported. Please save the file as .xlsx and upload again.")

    workbook = load_workbook(file_path, data_only=True, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(min_row=1, max_row=12, values_only=True))
    header_row_index, header_map = detect_header_row(rows, {"skucode", "sku"})

    if not header_row_index:
        raise ValueError("The item master file is missing a recognizable SKU Code header row.")

    items = []
    seen_skus = set()

    for row in sheet.iter_rows(min_row=header_row_index + 1, values_only=True):
        sku_code = first_present(row, header_map, "Sku Code", "SKU")
        normalized_sku = normalize_sku(sku_code)
        if not normalized_sku or normalized_sku in seen_skus:
            continue

        seen_skus.add(normalized_sku)
        items.append(
            {
                "skuCode": sku_code,
                "itemName": first_present(row, header_map, "Item Name"),
                "category": first_present(row, header_map, "Category"),
                "color": first_present(row, header_map, "Color", "Colour"),
                "brand": first_present(row, header_map, "Brand"),
                "hsnCode": first_present(row, header_map, "HSN Code"),
                "tat": first_present(row, header_map, "Tat"),
                "size": first_present(row, header_map, "Size"),
                "weight": first_present(row, header_map, "Weight"),
                "costPrice": first_number(row, header_map, "Cost Price"),
                "mrp": first_number(row, header_map, "MRP"),
                "batchGroup": first_present(row, header_map, "Batch Group"),
                "ean": first_present(row, header_map, "EAN"),
                "dimensions": first_present(row, header_map, "Dimensions"),
                "taxType": first_present(row, header_map, "Tax Type"),
                "enabled": first_present(row, header_map, "Enabled"),
                "itemType": first_present(row, header_map, "Type", "Item Type"),
                "expirable": first_present(row, header_map, "Expirable"),
                "skuType": first_present(row, header_map, "Sku Type"),
                "image": first_present(row, header_map, "Image"),
                "pageUrl": first_present(row, header_map, "Page URL"),
            }
        )

    return {"items": items}


def parse_pick_item(serial_no, sku_code, detail_lines):
    cleaned_lines = [clean_text(line) for line in detail_lines if clean_text(line)]
    summary_index = None

    for index in range(len(cleaned_lines) - 1, -1, -1):
        line = cleaned_lines[index]
        if re.search(r"\bDEFAULT\b", line, re.IGNORECASE) and re.search(r"\s\d+\s*$", line):
            summary_index = index
            break

    if summary_index is None and cleaned_lines:
        summary_index = len(cleaned_lines) - 1

    summary_line = cleaned_lines[summary_index] if summary_index is not None else ""
    name_lines = cleaned_lines[:summary_index] if summary_index is not None else cleaned_lines

    match = re.match(
        r"^(?:(?P<leading>.+?)\s+)?(?P<shelf>[A-Z0-9_-]+)\s+(?P<size>[A-Z0-9./+-]+)\s+(?P<color>.+?)\s+(?P<qty>\d+)$",
        summary_line,
        re.IGNORECASE,
    )

    leading = ""
    shelf_code = ""
    size = ""
    color = ""
    qty = 1

    if match:
        leading = clean_text(match.group("leading") or "")
        shelf_code = clean_text(match.group("shelf") or "")
        size = clean_text(match.group("size") or "")
        color = clean_text(match.group("color") or "")
        qty = int(match.group("qty") or 1)
    elif summary_line:
        parts = summary_line.rsplit(" ", 1)
        if len(parts) == 2 and parts[1].isdigit():
            qty = int(parts[1])
            leading = parts[0].strip()
        else:
            leading = summary_line.strip()

    name_parts = [part for part in name_lines if part]
    if leading:
        name_parts.append(leading)

    return {
        "serialNo": serial_no,
        "skuCode": sku_code,
        "pickListName": " ".join(name_parts).strip() or sku_code,
        "shelfCode": shelf_code,
        "size": size,
        "color": color,
        "qty": qty,
    }


def parse_pick_list_pdf(file_path):
    reader = PdfReader(file_path)
    full_text = "\n".join((page.extract_text() or "") for page in reader.pages)
    raw_lines = [line.rstrip() for line in full_text.splitlines()]

    # Accept several header label variants: "Picklist No:", "Pick List No:", "PL No:", etc.
    picklist_no_match = re.search(
        r"(?:Pick\s*List|Picklist|PL)\s*(?:No|#|Number)[.:\s]+([A-Za-z0-9_-]+)",
        full_text,
        re.IGNORECASE,
    )
    # Also try a bare PK-style number if the above doesn't match
    if not picklist_no_match:
        picklist_no_match = re.search(r"\b(PK[0-9]{3,})\b", full_text, re.IGNORECASE)

    created_match = re.search(
        r"Created[:\s]+([0-9]{1,2}\s+[A-Za-z]{3}\s+[0-9]{4}\s+[0-9]{2}:[0-9]{2})",
        full_text,
    )

    # Find the item list section (case-insensitive). Fall back to start of doc if not found.
    start_line = 0
    for idx, raw_line in enumerate(raw_lines):
        if re.search(r"pick\s+these\s+items", raw_line, re.IGNORECASE):
            start_line = idx + 1
            break

    items = []
    current_serial = None
    current_sku = ""
    detail_lines = []

    def flush_current():
        nonlocal current_serial, current_sku, detail_lines
        if current_serial is None or not current_sku:
            current_serial = None
            current_sku = ""
            detail_lines = []
            return
        items.append(parse_pick_item(current_serial, current_sku, detail_lines))
        current_serial = None
        current_sku = ""
        detail_lines = []

    for raw_line in raw_lines[start_line:]:
        line = clean_text(raw_line)
        if not line:
            continue

        if re.search(r"powered\s+by", line, re.IGNORECASE):
            break

        # Skip common PDF footer / header noise that starts with a number but is not
        # a pick-list serial: "Page 1 of 3", "1 of 3", standalone page numbers after
        # the item section ends, etc.
        if re.search(r"\bpage\b", line, re.IGNORECASE):
            continue

        # ── Format 1 ──────────────────────────────────────────────────────────
        # A line that is purely digits (optionally followed by a period or
        # closing paren) is an isolated serial number.
        # Examples: "1", "1.", "12", "42."
        if re.fullmatch(r"\d+[.)]?", line):
            flush_current()
            current_serial = int(re.sub(r"\D", "", line))
            continue

        # ── Format 2 ──────────────────────────────────────────────────────────
        # pypdf often merges the serial number and SKU onto a single line when
        # the PDF uses a table or two-column layout.
        # Examples:
        #   "1.PROD-SKU-001"          → serial=1, sku="PROD-SKU-001"
        #   "1. PROD-SKU-001"         → serial=1, sku="PROD-SKU-001"
        #   "1 PROD-SKU-001"          → serial=1, sku="PROD-SKU-001"
        #   "1 PROD-SKU-001 Name …"   → serial=1, sku="PROD-SKU-001", remainder=detail
        # Guard: SKU token must be ≥3 chars and contain no spaces (real SKU codes
        # never contain spaces; this rejects "1 of", "1 May", etc.).
        inline_match = re.match(
            r"^(\d{1,5})[.)\s]\s*([A-Za-z0-9][A-Za-z0-9_./-]{2,})(?:\s+(.+))?$",
            line,
        )
        if inline_match and int(inline_match.group(1)) > 0:
            flush_current()
            current_serial = int(inline_match.group(1))
            current_sku = clean_text(inline_match.group(2))
            remainder = clean_text(inline_match.group(3) or "")
            if remainder:
                detail_lines.append(remainder)
            continue

        if current_serial is None:
            continue

        if not current_sku:
            current_sku = line
            continue

        detail_lines.append(line)

    flush_current()

    return {
        "pickListNo": picklist_no_match.group(1) if picklist_no_match else Path(file_path).stem,
        "pickListCreatedAt": created_match.group(1) if created_match else None,
        "items": items,
    }


def extract_picklist_meta_from_rows(rows, file_path):
    values = []
    for row in rows[:12]:
        for cell in row:
            text = clean_text(cell)
            if text:
                values.append(text)

    joined = " | ".join(values)

    picklist_no = Path(file_path).stem
    created_at = None

    patterns = [
        r"Pick\s*List\s*No[:\s-]*([A-Za-z0-9-]+)",
        r"Picklist\s*No[:\s-]*([A-Za-z0-9-]+)",
        r"\b(PK[0-9]{3,})\b",
    ]
    for pattern in patterns:
        match = re.search(pattern, joined, re.IGNORECASE)
        if match:
            picklist_no = clean_text(match.group(1))
            break

    created_match = re.search(
        r"Created[:\s-]*([0-9]{1,2}[-/][0-9]{1,2}[-/][0-9]{2,4}(?:\s+[0-9]{1,2}:[0-9]{2})?)",
        joined,
        re.IGNORECASE,
    )
    if created_match:
        created_at = clean_text(created_match.group(1))

    return picklist_no, created_at


def parse_pick_list_excel(file_path):
    workbook = load_workbook(file_path, data_only=True, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    sample_rows = list(sheet.iter_rows(min_row=1, max_row=20, values_only=True))
    header_row_index, header_map = detect_header_row(sample_rows, {"skucode", "sku", "qty", "quantity"})

    if not header_row_index:
        raise ValueError("The pick list Excel file is missing a recognizable SKU/Qty header row.")

    pick_list_no, created_at = extract_picklist_meta_from_rows(sample_rows, file_path)
    items = []
    blank_streak = 0
    next_serial_no = 1

    for row in sheet.iter_rows(min_row=header_row_index + 1, values_only=True):
        sku_code = first_present(row, header_map, "Sku Code", "SKU")
        name = first_present(row, header_map, "Sku Name", "Item Name", "Product Name", "Name")
        shelf_code = first_present(row, header_map, "Shelf Code", "Shelf", "Bin Location")
        size = first_present(row, header_map, "Size")
        color = first_present(row, header_map, "Color", "Colour")
        qty = int(round(first_number(row, header_map, "Qty", "Quantity", "Pick Qty"))) or 0
        serial_text = first_present(row, header_map, "SI.No", "Sl.No", "Serial No", "S.No")

        if not sku_code and not name and qty == 0:
            blank_streak += 1
            if blank_streak >= 8 and items:
                break
            continue

        blank_streak = 0
        if not sku_code:
            continue

        try:
            serial_no = int(float(serial_text))
        except (TypeError, ValueError):
            serial_no = next_serial_no

        items.append(
            {
                "serialNo": serial_no,
                "skuCode": sku_code,
                "pickListName": name or sku_code,
                "shelfCode": shelf_code,
                "size": size,
                "color": color,
                "qty": max(1, qty),
            }
        )
        next_serial_no = serial_no + 1

    return {
        "pickListNo": pick_list_no,
        "pickListCreatedAt": created_at,
        "items": items,
    }


def parse_pick_list(file_path):
    suffix = Path(file_path).suffix.lower()
    if suffix == ".pdf":
        return parse_pick_list_pdf(file_path)
    if suffix == ".xls":
        raise ValueError("Legacy .xls pick list files are not supported. Please save the file as .xlsx and upload again.")
    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return parse_pick_list_excel(file_path)
    raise ValueError("Unsupported pick list file type. Use PDF or Excel.")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("mode", choices=["item-master", "pick-list"])
    parser.add_argument("--file", required=True)
    args = parser.parse_args()

    if args.mode == "item-master":
        result = parse_item_master(args.file)
    else:
        result = parse_pick_list(args.file)

    print(json.dumps(result))


if __name__ == "__main__":
    main()
